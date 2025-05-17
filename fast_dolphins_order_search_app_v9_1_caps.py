
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

file_path = "fastdolphins_merchandise-orders_250424141340.xlsx"
df = pd.read_excel(file_path)

# Ensure all required columns exist
for col in ["Complete", "Swim Cap", "Cap Received"]:
    if col not in df.columns:
        df[col] = ""

df.fillna("", inplace=True)
df["family_id"] = df["parent1_first_name"] + " " + df["parent1_last_name"] + " - " + df["athletes"]

st.markdown("""
<style>
.main {background-color: #F0F8FF;}
.title {color: #003366; font-size: 32px; font-weight: bold;}
.subtitle {color: #003366; font-size: 22px; font-weight: bold;}
.order {font-size: 20px;}
.warning {color: red; font-weight: bold;}
</style>
""", unsafe_allow_html=True)

st.markdown("<div class='title'>FAST Dolphins Orders</div>", unsafe_allow_html=True)

if "return_to_search" not in st.session_state:
    st.session_state.return_to_search = False

view_completed = st.toggle("Show Completed Orders Only")

if view_completed:
    completed_orders = df[df["Complete"] == "Yes"]
    if completed_orders.empty:
        st.write("✅ No completed orders yet.")
    else:
        st.markdown("<div class='subtitle'>✅ Completed Orders:</div>", unsafe_allow_html=True)
        for _, row in completed_orders.iterrows():
            info = f"{row['family_id']} — {row['option_name']} — Qty {row['quantity']}"
            if row["Cap Received"] == "Yes":
                info += " 🧢 Cap Received"
            st.markdown(f"<div class='order'>🟦 {info}</div>", unsafe_allow_html=True)
else:
    st.write("Search by athlete or parent name to view shirt sizes and mark complete.")

    if st.session_state.return_to_search:
        st.session_state.return_to_search = False
        st.rerun()

    search_term = st.text_input("Enter name (athlete or parent):")

    if search_term:
        results = df[
            df['athletes'].str.contains(search_term, case=False, na=False) |
            df['for_athlete'].str.contains(search_term, case=False, na=False) |
            df['parent1_first_name'].str.contains(search_term, case=False, na=False) |
            df['parent1_last_name'].str.contains(search_term, case=False, na=False)
        ].copy()

        if results.empty:
            st.write("No orders found for this person.")
        else:
            st.markdown("<div class='subtitle'>Shirt Orders:</div>", unsafe_allow_html=True)
            results["Mark Complete"] = False
            results["Cap Given"] = False

            for i in results.index:
                family = results.at[i, "family_id"]
                size = results.at[i, "option_name"]
                qty = results.at[i, "quantity"]
                cap = results.at[i, "Swim Cap"]
                already_complete = results.at[i, "Complete"] == "Yes"

                if already_complete:
                    cap_note = "🧢 Cap Received" if results.at[i, "Cap Received"] == "Yes" else ""
                    st.markdown(f"<div class='order'>🟦 {family} — {size} — Qty {qty} <span class='warning'>(ALREADY RECEIVED)</span> {cap_note}</div>", unsafe_allow_html=True)
                else:
                    results.at[i, "Mark Complete"] = st.checkbox(f"{family} — {size} — Qty {qty}", key=f"shirt_{i}")
                    if cap.lower() == "yes":
                        results.at[i, "Cap Given"] = st.checkbox(f"🧢 Give cap to {family}", key=f"cap_{i}")

            if st.button("Mark Selected as Complete"):
                for i in results.index:
                    if results.at[i, "Mark Complete"]:
                        df.at[i, "Complete"] = "Yes"
                    if results.at[i, "Cap Given"]:
                        df.at[i, "Cap Received"] = "Yes"

                df.to_excel(file_path, index=False)

                wb = load_workbook(file_path)
                ws = wb.active
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    complete_cell = row[len(df.columns) - 3]
                    if complete_cell.value == "Yes":
                        for cell in row:
                            cell.fill = green_fill
                wb.save(file_path)

                st.success("✅ Orders updated successfully.")
                st.session_state.return_to_search = True
                st.rerun()
    else:
        st.write("Please enter a name to search.")
