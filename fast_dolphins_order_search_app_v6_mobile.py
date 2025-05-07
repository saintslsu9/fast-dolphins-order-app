
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# File path
file_path = "fastdolphins_merchandise-orders_250424141340.xlsx"

# Load data
df = pd.read_excel(file_path)

# Add 'Complete' column if it doesn't exist
if "Complete" not in df.columns:
    df["Complete"] = ""

# Fill NaN values for cleaner display
df["for_athlete"] = df["for_athlete"].fillna("")
df["athletes"] = df["athletes"].fillna("")
df["parent1_first_name"] = df["parent1_first_name"].fillna("")
df["parent1_last_name"] = df["parent1_last_name"].fillna("")

# Build a unique identifier for families -> Parent full name + athlete
df["family_id"] = df["parent1_first_name"] + " " + df["parent1_last_name"] + " - " + df["athletes"]

# Page styling optimized for mobile (larger fonts, clean layout)
st.markdown(
    \"\"\"
    <style>
    .main {background-color: #F0F8FF;}
    .title {color: #003366; font-size: 32px; font-weight: bold;}
    .subtitle {color: #003366; font-size: 22px; font-weight: bold;}
    .order {font-size: 20px;}
    .warning {color: red; font-weight: bold;}
    </style>
    \"\"\", unsafe_allow_html=True
)

st.markdown("<div class='title'>FAST Dolphins Orders</div>", unsafe_allow_html=True)

# Initialize session state to control search view after completion
if "return_to_search" not in st.session_state:
    st.session_state.return_to_search = False

# Option to view completed orders
view_completed = st.toggle("Show Completed Orders Only")

if view_completed:
    completed_orders = df[df["Complete"] == "Yes"]
    if completed_orders.empty:
        st.write("✅ No completed orders yet.")
    else:
        st.markdown("<div class='subtitle'>✅ Completed Orders:</div>", unsafe_allow_html=True)
        for i in completed_orders.index:
            family = completed_orders.at[i, "family_id"]
            size = completed_orders.at[i, "option_name"]
            qty = completed_orders.at[i, "quantity"]
            st.markdown(f"<div class='order'>🟦 {family} — {size} — Qty {qty}</div>", unsafe_allow_html=True)
else:
    st.write("Search by athlete or parent name to view shirt sizes and mark complete.")

    # Reset search view if coming back from completion action
    if st.session_state.return_to_search:
        st.session_state.return_to_search = False
        st.experimental_rerun()

    search_term = st.text_input("Enter name (athlete or parent):")

    if search_term:
        results = df[
            df['athletes'].str.contains(search_term, case=False, na=False) |
            df['for_athlete'].str.contains(search_term, case=False, na=False) |
            df['parent1_first_name'].str.contains(search_term, case=False, na=False) |
            df['parent1_last_name'].str.contains(search_term, case=False, na=False)
        ].copy()

        if results.empty:
            st.write("No orders found for this person.")
        else:
            st.markdown("<div class='subtitle'>Shirt Orders:</div>", unsafe_allow_html=True)
            results["Mark Complete"] = False

            for i in results.index:
                family = results.at[i, "family_id"]
                size = results.at[i, "option_name"]
                qty = results.at[i, "quantity"]
                already_complete = results.at[i, "Complete"] == "Yes"

                if already_complete:
                    st.markdown(f"<div class='order'>🟦 {family} — {size} — Qty {qty} <span class='warning'>(ALREADY RECEIVED)</span></div>", unsafe_allow_html=True)
                else:
                    results.at[i, "Mark Complete"] = st.checkbox(
                        f"{family} — {size} — Qty {qty}",
                        key=i
                    )

            if st.button("Mark Selected as Complete"):
                for i in results.index:
                    if results.at[i, "Mark Complete"]:
                        df.at[i, "Complete"] = "Yes"

                df.to_excel(file_path, index=False)

                wb = load_workbook(file_path)
                ws = wb.active

                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    complete_cell = row[len(df.columns) - 1]
                    if complete_cell.value == "Yes":
                        for cell in row:
                            cell.fill = green_fill

                wb.save(file_path)
                st.success("✅ Selected orders marked complete and Excel file updated!")

                st.session_state.return_to_search = True
                st.experimental_rerun()
    else:
        st.write("Please enter a name to search.")
"""

# Save the V9 app file
v9_app_file_path = "/mnt/data/fast_dolphins_order_search_app_v9_mobile.py"
with open(v9_app_file_path, "w") as f:
    f.write(v9_web_app_code)

# Create ZIP for V9
import zipfile
zip_path_v9 = "/mnt/data/FAST_Dolphins_Order_Search_App_V9_Mobile.zip"
with zipfile.ZipFile(zip_path_v9, 'w') as zipf:
    zipf.write(v9_app_file_path, arcname="fast_dolphins_order_search_app_v9_mobile.py")
